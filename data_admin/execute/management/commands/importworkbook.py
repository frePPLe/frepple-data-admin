from datetime import datetime
import logging
from openpyxl import load_workbook

from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_permission_codename
from django.contrib.auth.models import Group
from django.contrib.contenttypes.models import ContentType
from django.db import DEFAULT_DB_ALIAS
from django.db import transaction
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from django.utils.text import capfirst
from django.utils.encoding import force_text
from django.template.loader import render_to_string

from .... import __version__
from ....common.middleware import _thread_locals
from ....common.models import User, Comment
from ....common.report import GridReport, matchesModelName
from ....common.dataload import parseExcelWorksheet
from ...models import Task


logger = logging.getLogger(__name__)

# A list of models with some special, administrative purpose.
# They should be excluded from bulk import, export and erasing actions.
EXCLUDE_FROM_BULK_OPERATIONS = (Group, User, Comment)


class Command(BaseCommand):

    # help = "Loads XLS workbook file into the frePPLe database"
    help = "command not implemented yet"

    requires_system_checks = False

    def add_arguments(self, parser):
        parser.add_argument("--user", help="User running the command")
        parser.add_argument(
            "--database",
            default=DEFAULT_DB_ALIAS,
            help="Nominates a specific database to load data from and export results into",
        )
        parser.add_argument(
            "--task",
            type=int,
            help="Task identifier (generated automatically if not provided)",
        )
        parser.add_argument("file", nargs="+", help="workbook file name")

    def get_version(self):
        return __version__

    def handle(self, **options):
        # Pick up the options
        now = datetime.now()
        self.database = options["database"]
        if self.database not in settings.DATABASES:
            raise CommandError("No database settings known for '%s'" % self.database)
        if options["user"]:
            try:
                self.user = (
                    User.objects.all()
                    .using(self.database)
                    .get(username=options["user"])
                )
            except Exception:
                raise CommandError("User '%s' not found" % options["user"])
        else:
            self.user = None
        timestamp = now.strftime("%Y%m%d%H%M%S")
        if self.database == DEFAULT_DB_ALIAS:
            logfile = "importworkbook-%s.log" % timestamp
        else:
            logfile = "importworkbook_%s-%s.log" % (self.database, timestamp)

        task = None
        try:
            setattr(_thread_locals, "database", self.database)
            # Initialize the task
            if options["task"]:
                try:
                    task = (
                        Task.objects.all().using(self.database).get(pk=options["task"])
                    )
                except Exception:
                    raise CommandError("Task identifier not found")
                if (
                    task.started
                    or task.finished
                    or task.status != "Waiting"
                    or task.name not in ("frepple_importworkbook", "importworkbook")
                ):
                    raise CommandError("Invalid task identifier")
                task.status = "0%"
                task.started = now
            else:
                task = Task(
                    name="importworkbook",
                    submitted=now,
                    started=now,
                    status="0%",
                    user=self.user,
                )
            task.arguments = " ".join(options["file"])
            task.save(using=self.database)

            all_models = [
                (ct.model_class(), ct.pk)
                for ct in ContentType.objects.all()
                if ct.model_class()
            ]
            try:
                with transaction.atomic(using=self.database):
                    # Find all models in the workbook
                    if "filename" not in locals():
                        filename = options["file"]
                    for file in filename:
                        wb = load_workbook(
                            filename=file, read_only=True, data_only=True
                        )
                        models = []
                        for ws_name in wb.sheetnames:
                            # Find the model
                            model = None
                            contenttype_id = None
                            for m, ct in all_models:
                                if matchesModelName(ws_name, m):
                                    model = m
                                    contenttype_id = ct
                                    break
                            if not model or model in EXCLUDE_FROM_BULK_OPERATIONS:
                                print(
                                    force_text(
                                        _("Ignoring data in worksheet: %s") % ws_name
                                    )
                                )
                                # yield '<div class="alert alert-warning">' + force_text(_("Ignoring data in worksheet: %s") % ws_name) + '</div>'
                            elif not self.user.has_perm(
                                "%s.%s"
                                % (
                                    model._meta.app_label,
                                    get_permission_codename("add", model._meta),
                                )
                            ):
                                # Check permissions
                                print(
                                    force_text(
                                        _("You don't permissions to add: %s") % ws_name
                                    )
                                )
                                # yield '<div class="alert alert-danger">' + force_text(_("You don't permissions to add: %s") % ws_name) + '</div>'
                            else:
                                deps = set([model])
                                GridReport.dependent_models(model, deps)
                                models.append((ws_name, model, contenttype_id, deps))

                        # Sort the list of models, based on dependencies between models
                        models = GridReport.sort_models(models)

                        # Process all rows in each worksheet
                        for ws_name, model, contenttype_id, dependencies in models:
                            print(
                                force_text(
                                    _("Processing data in worksheet: %s") % ws_name
                                )
                            )
                            # yield '<strong>' + force_text(_("Processing data in worksheet: %s") % ws_name) + '</strong><br>'
                            # yield ('<div class="table-responsive">'
                            # '<table class="table table-condensed" style="white-space: nowrap;"><tbody>')
                            numerrors = 0
                            numwarnings = 0
                            firsterror = True
                            ws = wb[ws_name]
                            for error in parseExcelWorksheet(
                                model,
                                ws,
                                user=self.user,
                                database=self.database,
                                ping=True,
                            ):
                                if error[0] == logging.DEBUG:
                                    # Yield some result so we can detect disconnect clients and interrupt the upload
                                    # yield ' '
                                    continue
                                if firsterror and error[0] in (
                                    logging.ERROR,
                                    logging.WARNING,
                                ):
                                    print(
                                        "%s %s %s %s %s%s%s"
                                        % (
                                            capfirst(_("worksheet")),
                                            capfirst(_("row")),
                                            capfirst(_("field")),
                                            capfirst(_("value")),
                                            capfirst(_("error")),
                                            " / ",
                                            capfirst(_("warning")),
                                        )
                                    )
                                    # yield '<tr><th class="sr-only">%s</th><th>%s</th><th>%s</th><th>%s</th><th>%s%s%s</th></tr>' % (
                                    #   capfirst(_("worksheet")), capfirst(_("row")),
                                    #   capfirst(_("field")), capfirst(_("value")),
                                    #   capfirst(_("error")), " / ", capfirst(_("warning"))
                                    #   )
                                    firsterror = False
                                if error[0] == logging.ERROR:
                                    print(
                                        "%s %s %s %s %s: %s"
                                        % (
                                            ws_name,
                                            error[1] if error[1] else "",
                                            error[2] if error[2] else "",
                                            error[3] if error[3] else "",
                                            capfirst(_("error")),
                                            error[4],
                                        )
                                    )
                                    # yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                    #   ws_name,
                                    #   error[1] if error[1] else '',
                                    #   error[2] if error[2] else '',
                                    #   error[3] if error[3] else '',
                                    #   capfirst(_('error')),
                                    #   error[4]
                                    #   )
                                    numerrors += 1
                                elif error[1] == logging.WARNING:
                                    print(
                                        "%s %s %s %s %s: %s"
                                        % (
                                            ws_name,
                                            error[1] if error[1] else "",
                                            error[2] if error[2] else "",
                                            error[3] if error[3] else "",
                                            capfirst(_("warning")),
                                            error[4],
                                        )
                                    )
                                    # yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                    #   ws_name,
                                    #   error[1] if error[1] else '',
                                    #   error[2] if error[2] else '',
                                    #   error[3] if error[3] else '',
                                    #   capfirst(_('warning')),
                                    #   error[4]
                                    #   )
                                    numwarnings += 1
                                else:
                                    print(
                                        "%s %s %s %s %s %s"
                                        % (
                                            "danger" if numerrors > 0 else "success",
                                            ws_name,
                                            error[1] if error[1] else "",
                                            error[2] if error[2] else "",
                                            error[3] if error[3] else "",
                                            error[4],
                                        )
                                    )
                            #     yield '<tr class=%s><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                            #       "danger" if numerrors > 0 else 'success',
                            #       ws_name,
                            #       error[1] if error[1] else '',
                            #       error[2] if error[2] else '',
                            #       error[3] if error[3] else '',
                            #       error[4]
                            #       )
                            # yield '</tbody></table></div>'
                        print("%s" % _("Done"))
                        # yield '<div><strong>%s</strong></div>' % _("Done")
            except GeneratorExit:
                logger.warning("Connection Aborted")
        except Exception as e:
            if task:
                task.status = "Failed"
                task.message = "%s" % e
                task.finished = datetime.now()
            raise e

        finally:
            setattr(_thread_locals, "database", None)
            if task:
                task.save(using=self.database)

        # Task update
        task.status = "Done"
        task.finished = datetime.now()
        task.processid = None
        task.save(using=self.database, update_fields=["status", "finished"])

        return _("Done")

    # accordion template
    title = _("Import a spreadsheet")
    index = 1100
    help_url = "command-reference.html#importworkbook"

    @staticmethod
    def getHTML(request):
        return render_to_string("commands/importworkbook.html", request=request)

import codecs
from datetime import datetime
from time import localtime, strftime
import csv
import gzip
from openpyxl import load_workbook
import os
import logging

from django.conf import settings
from django.contrib.auth import get_permission_codename
from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import connections, DEFAULT_DB_ALIAS, transaction
from django.template.loader import render_to_string
from django.utils import translation
from django.utils.formats import get_format
from django.utils.translation import gettext_lazy as _

from ...models import Task
from ....common.middleware import _thread_locals
from ....common.report import GridReport, matchesModelName
from .... import __version__
from ....common.dataload import parseCSVdata, parseExcelWorksheet
from ....common.models import User, NotificationFactory
from ....common.report import EXCLUDE_FROM_BULK_OPERATIONS, create_connection

logger = logging.getLogger(__name__)


class Command(BaseCommand):

    help = """
    Loads CSV files from the configured FILEUPLOADFOLDER folder into the frePPLe database.
    The data files should have the extension .csv or .csv.gz, and the file name should
    start with the name of the data model.
    """

    requires_system_checks = False

    def add_arguments(self, parser):
        parser.add_argument("--user", help="User running the command")
        parser.add_argument(
            "--database",
            default=DEFAULT_DB_ALIAS,
            help="Nominates a specific database to load the data into",
        )
        parser.add_argument(
            "--task",
            type=int,
            help="Task identifier (generated automatically if not provided)",
        )

    def get_version(self):
        return __version__

    def handle(self, **options):
        # Pick up the options
        now = datetime.now()
        self.database = options["database"]
        if self.database not in settings.DATABASES:
            raise CommandError("No database settings known for '%s'" % self.database)
        if options["user"]:
            try:
                self.user = (
                    User.objects.all()
                    .using(self.database)
                    .get(username=options["user"])
                )
            except Exception:
                raise CommandError("User '%s' not found" % options["user"])
        else:
            self.user = None
        timestamp = now.strftime("%Y%m%d%H%M%S")
        if self.database == DEFAULT_DB_ALIAS:
            logfile = "importfromfolder-%s.log" % timestamp
        else:
            logfile = "importfromfolder_%s-%s.log" % (self.database, timestamp)

        try:
            handler = logging.FileHandler(
                os.path.join(settings.FREPPLE_LOGDIR, logfile), encoding="utf-8"
            )
            # handler.setFormatter(logging.Formatter(settings.LOGGING['formatters']['simple']['format']))
            logger.addHandler(handler)
            logger.propagate = False
        except Exception as e:
            print("%s Failed to open logfile %s: %s" % (datetime.now(), logfile, e))

        task = None
        errors = [0, 0]
        try:
            setattr(_thread_locals, "database", self.database)
            # Initialize the task
            if options["task"]:
                try:
                    task = (
                        Task.objects.all().using(self.database).get(pk=options["task"])
                    )
                except Exception:
                    raise CommandError("Task identifier not found")
                if (
                    task.started
                    or task.finished
                    or task.status != "Waiting"
                    or task.name not in ("frepple_importfromfolder", "importfromfolder")
                ):
                    raise CommandError("Invalid task identifier")
                task.status = "0%"
                task.started = now
                task.logfile = logfile
            else:
                task = Task(
                    name="importfromfolder",
                    submitted=now,
                    started=now,
                    status="0%",
                    user=self.user,
                    logfile=logfile,
                )
            task.processid = os.getpid()
            task.save(using=self.database)

            # Choose the right self.delimiter and language
            self.delimiter = (
                get_format("DECIMAL_SEPARATOR", settings.LANGUAGE_CODE, True) == ","
                and ";"
                or ","
            )
            translation.activate(settings.LANGUAGE_CODE)
            self.SQLrole = settings.DATABASES[self.database].get(
                "SQL_ROLE", "report_role"
            )

            # Execute
            if "FILEUPLOADFOLDER" in settings.DATABASES[
                self.database
            ] and os.path.isdir(settings.DATABASES[self.database]["FILEUPLOADFOLDER"]):

                # Open the logfile
                logger.info(
                    "%s Started importfromfolder\n"
                    % datetime.now().replace(microsecond=0)
                )

                all_models = [
                    (ct.model_class(), ct.pk)
                    for ct in ContentType.objects.all()
                    if ct.model_class()
                ]
                models = []
                for ifile in os.listdir(
                    settings.DATABASES[self.database]["FILEUPLOADFOLDER"]
                ):
                    if not ifile.lower().endswith(
                        (
                            ".sql",
                            ".sql.gz",
                            ".csv",
                            ".csv.gz",
                            ".cpy",
                            ".cpy.gz",
                            ".xlsx",
                        )
                    ):
                        continue
                    filename0 = ifile.split(".")[0].split(" (")[0]

                    model = None
                    contenttype_id = None
                    for m, ct in all_models:
                        if matchesModelName(filename0, m):
                            model = m
                            contenttype_id = ct
                            break

                    if not model or model in EXCLUDE_FROM_BULK_OPERATIONS:
                        logger.info(
                            "%s Ignoring data in file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                    elif self.user and not self.user.has_perm(
                        "%s.%s"
                        % (
                            model._meta.app_label,
                            get_permission_codename("add", model._meta),
                        )
                    ):
                        # Check permissions
                        logger.info(
                            "%s You don't have permissions to add: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                    else:
                        deps = set([model])
                        GridReport.dependent_models(model, deps)

                        models.append((ifile, model, contenttype_id, deps))

                # Sort the list of models, based on dependencies between models
                models = GridReport.sort_models(models)

                i = 0
                cnt = len(models)
                for ifile, model, contenttype_id, dependencies in models:
                    task.status = str(int(10 + i / cnt * 80)) + "%"
                    task.message = "Processing data file %s" % ifile
                    task.save(using=self.database)
                    i += 1
                    filetoparse = os.path.join(
                        os.path.abspath(
                            settings.DATABASES[self.database]["FILEUPLOADFOLDER"]
                        ),
                        ifile,
                    )
                    if ifile.lower().endswith((".sql", ".sql.gz")):
                        logger.info(
                            "%s Started executing SQL statements from file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                        errors[0] += self.executeSQLfile(filetoparse)
                        logger.info(
                            "%s Finished executing SQL statements from file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                    elif ifile.lower().endswith((".cpy", ".cpy.gz")):
                        logger.info(
                            "%s Started uploading copy file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                        errors[0] += self.executeCOPYfile(model, filetoparse)
                        logger.info(
                            "%s Finished uploading copy file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                    elif ifile.lower().endswith(".xlsx"):
                        logger.info(
                            "%s Started processing data in Excel file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                        returnederrors = self.loadExcelfile(model, filetoparse)
                        errors[0] += returnederrors[0]
                        errors[1] += returnederrors[1]
                        logger.info(
                            "%s Finished processing data in file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                    else:
                        logger.info(
                            "%s Started processing data in CSV file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
                        returnederrors = self.loadCSVfile(model, filetoparse)
                        errors[0] += returnederrors[0]
                        errors[1] += returnederrors[1]
                        logger.info(
                            "%s Finished processing data in CSV file: %s"
                            % (datetime.now().replace(microsecond=0), ifile)
                        )
            else:
                errors[0] += 1
                cnt = 0
                logger.error(
                    "%s Failed, folder does not exist"
                    % datetime.now().replace(microsecond=0)
                )

            # Task update
            if errors[0] > 0:
                task.status = "Failed"
                if not cnt:
                    task.message = "Destination folder does not exist"
                else:
                    task.message = (
                        "Uploaded %s data files with %s errors and %s warnings"
                        % (cnt, errors[0], errors[1])
                    )
            else:
                task.status = "Done"
                task.message = "Uploaded %s data files with %s warnings" % (
                    cnt,
                    errors[1],
                )
            task.finished = datetime.now()

        except KeyboardInterrupt:
            if task:
                task.status = "Cancelled"
                task.message = "Cancelled"
            logger.info("%s Cancelled\n" % datetime.now().replace(microsecond=0))

        except Exception as e:
            logger.error("%s Failed" % datetime.now().replace(microsecond=0))
            if task:
                task.status = "Failed"
                task.message = "%s" % e
            raise e

        finally:
            setattr(_thread_locals, "database", None)
            if task:
                if errors[0] == 0:
                    task.status = "Done"
                else:
                    task.status = "Failed"
                task.processid = None
                task.finished = datetime.now()
                task.save(using=self.database)
            logger.info(
                "%s End of importfromfolder\n" % datetime.now().replace(microsecond=0)
            )

    def executeCOPYfile(self, model, ifile):
        """
        Use the copy command to upload data into the database
        The filename must be equal to the table name (E.g : demand, buffer, forecast...)
        The first line of the file must contain the columns to popualate, comma separated
        """
        cursor = connections[self.database].cursor()
        try:

            if ifile.lower().endswith(".gz"):
                file_open = gzip.open
            else:
                file_open = open

            # Retrieve the header line of the file
            tableName = model._meta.db_table
            f = file_open(ifile, "rt")
            firstLine = f.readline().rstrip()
            f.close()

            # Validate the data fields in the header
            headers = []
            for f in firstLine.split(","):
                col = f.strip().strip("#").strip('"').lower() if f else ""
                dbfield = None
                for i in model._meta.fields:
                    # Try with database field name
                    if col == i.get_attname():
                        dbfield = i.get_attname()
                        break
                    # Try with translated field names
                    elif col == i.name.lower() or col == i.verbose_name.lower():
                        dbfield = i.get_attname()
                        break
                    if translation.get_language() != "en":
                        # Try with English field names
                        with translation.override("en"):
                            if col == i.name.lower() or col == i.verbose_name.lower():
                                dbfield = i.get_attname()
                                break
                if dbfield:
                    headers.append('"%s"' % dbfield)
                else:
                    raise Exception("Invalid field name '%s'" % col)

            # count how many records in table before the copy operation
            cursor.execute("select count(*) from %s" % tableName)
            countBefore = cursor.fetchone()[0]

            # Load the data records
            copyFile = file_open(ifile)
            cursor.copy_expert(
                "copy %s (%s) from STDIN with delimiter ',' csv header"
                % (tableName, ",".join(headers)),
                copyFile,
            )

            # count how many records in table after the copy operation
            cursor.execute("select count(*) from %s" % tableName)
            countAfter = cursor.fetchone()[0]

            logger.info(
                "%s %s records uploaded into table %s"
                % (
                    datetime.now().replace(microsecond=0),
                    (countAfter - countBefore),
                    tableName,
                )
            )
            return 0

        except Exception as e:
            logger.error(
                "%s Error uploading COPY file: %s"
                % (datetime.now().replace(microsecond=0), e)
            )
            return 1
        finally:
            # Need to force closing the connection. Otherwise we keep the
            # connection in the restricted role.
            connections[self.database].close()

    def executeSQLfile(self, ifile):
        """
        Execute statements from a text with SQL statements.
        """
        if ifile.lower().endswith(".gz"):
            file_open = gzip.open
        else:
            file_open = open
        conn = None
        try:
            conn = create_connection(self.database)
            with conn.cursor() as cursor:
                if self.SQLrole:
                    cursor.execute("set role %s", (self.SQLrole,))
                cursor.execute(file_open(ifile, "rt").read())
            return 0
        except Exception as e:
            logger.error(
                "%s Error executing SQL: %s"
                % (datetime.now().replace(microsecond=0), e)
            )
            return 1
        finally:
            if conn:
                conn.close()

    def loadCSVfile(self, model, file):
        errorcount = 0
        warningcount = 0
        datafile = EncodedCSVReader(file, delimiter=self.delimiter)
        try:
            with transaction.atomic(using=self.database):
                for error in parseCSVdata(
                    model, datafile, user=self.user, database=self.database
                ):
                    if error[0] == logging.ERROR:
                        logger.error(
                            "%s Error: %s%s%s%s"
                            % (
                                datetime.now().replace(microsecond=0),
                                "Row %s: " % error[1] if error[1] else "",
                                "field %s: " % error[2] if error[2] else "",
                                "%s: " % error[3] if error[3] else "",
                                error[4],
                            )
                        )
                        errorcount += 1
                    elif error[0] == logging.WARNING:
                        logger.warning(
                            "%s Warning: %s%s%s%s"
                            % (
                                datetime.now().replace(microsecond=0),
                                "Row %s: " % error[1] if error[1] else "",
                                "field %s: " % error[2] if error[2] else "",
                                "%s: " % error[3] if error[3] else "",
                                error[4],
                            )
                        )
                        warningcount += 1
                    else:
                        logger.info(
                            "%s %s%s%s%s"
                            % (
                                datetime.now().replace(microsecond=0),
                                "Row %s: " % error[1] if error[1] else "",
                                "field %s: " % error[2] if error[2] else "",
                                "%s: " % error[3] if error[3] else "",
                                error[4],
                            )
                        )

            # Records are committed. Launch notification generator now.
            NotificationFactory.launchWorker(database=self.database, url=None)

        except Exception:
            errorcount += 1
            logger.error(
                "%s Error: Invalid data format - skipping the file \n"
                % datetime.now().replace(microsecond=0)
            )
        return [errorcount, warningcount]

    def loadExcelfile(self, model, file):
        errorcount = 0
        warningcount = 0
        try:
            with transaction.atomic(using=self.database):
                wb = load_workbook(filename=file, read_only=True, data_only=True)
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    for error in parseExcelWorksheet(
                        model, ws, user=self.user, database=self.database
                    ):
                        if error[0] == logging.ERROR:
                            logger.error(
                                "%s Error: %s%s%s%s"
                                % (
                                    datetime.now().replace(microsecond=0),
                                    "Row %s: " % error[1] if error[1] else "",
                                    "field %s: " % error[2] if error[2] else "",
                                    "%s: " % error[3] if error[3] else "",
                                    error[4],
                                )
                            )
                            errorcount += 1
                        elif error[0] == logging.WARNING:
                            logger.warning(
                                "%s Warning: %s%s%s%s"
                                % (
                                    datetime.now().replace(microsecond=0),
                                    "Row %s: " % error[1] if error[1] else "",
                                    "field %s: " % error[2] if error[2] else "",
                                    "%s: " % error[3] if error[3] else "",
                                    error[4],
                                )
                            )
                            warningcount += 1
                        else:
                            logger.info(
                                "%s %s%s%s%s"
                                % (
                                    datetime.now().replace(microsecond=0),
                                    "Row %s: " % error[1] if error[1] else "",
                                    "field %s: " % error[2] if error[2] else "",
                                    "%s: " % error[3] if error[3] else "",
                                    error[4],
                                )
                            )
            # Records are committed. Launch notification generator now.
            NotificationFactory.launchWorker(database=self.database, url=None)
        except Exception:
            errorcount += 1
            logger.error(
                "%s Error: Invalid data format - skipping the file \n"
                % datetime.now().replace(microsecond=0)
            )
        return [errorcount, warningcount]

    # accordion template
    title = _("Import data files")
    index = 1300
    help_url = "command-reference.html#importfromfolder"

    @staticmethod
    def getHTML(request):

        if (
            "FILEUPLOADFOLDER" not in settings.DATABASES[request.database]
            or not request.user.is_superuser
        ):
            return None

        # Function to convert from bytes to human readabl format
        def sizeof_fmt(num):
            for unit in ["", "Ki", "Mi", "Gi", "Ti", "Pi", "Ei", "Zi"]:
                if abs(num) < 1024.0:
                    return "%3.1f%sB" % (num, unit)
                num /= 1024.0
            return "%.1f%sB" % (num, "Yi")

        filestoupload = []
        if "FILEUPLOADFOLDER" in settings.DATABASES[request.database]:
            uploadfolder = settings.DATABASES[request.database]["FILEUPLOADFOLDER"]
            if os.path.isdir(uploadfolder):
                tzoffset = GridReport.getTimezoneOffset(request)
                for file in os.listdir(uploadfolder):
                    if file.endswith(
                        (
                            ".csv",
                            ".csv.gz",
                            ".xlsx",
                            ".cpy",
                            ".sql",
                            ".cpy.gz",
                            ".sql.gz",
                        )
                    ):
                        filestoupload.append(
                            [
                                file,
                                strftime(
                                    "%Y-%m-%d %H:%M:%S",
                                    localtime(
                                        os.stat(
                                            os.path.join(uploadfolder, file)
                                        ).st_mtime
                                        + tzoffset.total_seconds()
                                    ),
                                ),
                                sizeof_fmt(
                                    os.stat(os.path.join(uploadfolder, file)).st_size
                                ),
                            ]
                        )

        return render_to_string(
            "commands/importfromfolder.html",
            {"filestoupload": filestoupload},
            request=request,
        )


class EncodedCSVReader:
    """
    A CSV reader which will iterate over lines in the CSV data buffer.
    The reader will scan the BOM header in the data to detect the right encoding.
    """

    def __init__(self, datafile, **kwds):
        # Read the file into memory
        # TODO Huge file uploads can overwhelm your system!

        # Detect the encoding of the data by scanning the BOM.
        # Skip the BOM header if it is found.

        if datafile.lower().endswith(".gz"):
            file_open = gzip.open
        else:
            file_open = open
        self.reader = file_open(datafile, "rb")
        data = self.reader.read(5)
        self.reader.close()
        if data.startswith(codecs.BOM_UTF32_BE):
            self.reader = file_open(datafile, "rt", encoding="utf_32_be")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF32_LE):
            self.reader = file_open(datafile, "rt", encoding="utf_32_le")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF16_BE):
            self.reader = file_open(datafile, "rt", encoding="utf_16_be")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF16_LE):
            self.reader = file_open(datafile, "rt", encoding="utf_16_le")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF8):
            self.reader = file_open(datafile, "rt", encoding="utf_8")
            self.reader.read(1)
        else:
            # No BOM header found. We assume the data is encoded in the default CSV character set.
            self.reader = file_open(datafile, "rt", encoding=settings.CSV_CHARSET)

        # Open the file
        self.csvreader = csv.reader(self.reader, **kwds)

    def __next__(self):
        return next(self.csvreader)

    def __iter__(self):
        return self
